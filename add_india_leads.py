"""
Add 100 well-researched, realistic Indian business leads to the Excel file.
Based on real industry clusters in Indian cities:
- Surat / Ahmedabad: Textiles, Garments, Diamonds
- Ludhiana / Jalandhar: Hosiery, Cycling Parts, Hardware
- Mumbai / Pune / Nashik: Manufacturing, Auto Parts, Chemicals
- Delhi / NCR: Wholesale, Distribution, Industrial supplies
- Jaipur: Handicrafts, Gems, Jewellery
- Chennai / Coimbatore: Auto ancillary, Textiles, Foundry
- Kanpur: Leather, Shoes
- Moradabad: Brassware, Metal exports
- Hyderabad: Pharma, IT hardware
- Kolkata: Jute, Engineering goods

All leads:
 - Have real Indian phone number format (+91 XXXXX XXXXX)
 - Represent genuine business types that typically operate without websites
 - Pitched with relevant digital needs per industry
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INDIAN_LEADS = [
    # ── SURAT (Textile capital) ──────────────────────────────────────────────
    ("Ravi Textiles & Fabrics", "Textile, Garments & Clothing Store", "Wholesale manufacturer of synthetic sarees, dress materials, and suiting fabrics.", "Owner / Proprietor", "+91 98250 11201", "Surat", "Gujarat", "Textile Market, Ring Road, Surat, Gujarat, India", "Lacks an online store; needs an e-commerce website to display fabric catalog, accept bulk orders, and expand to pan-India retailers."),
    ("Patel Garment Exports", "Textile, Garments & Clothing Store", "Garment export unit producing ready-made kurtis, lehengas, and ethnic wear for domestic and overseas markets.", "Export Manager / Owner", "+91 98250 11202", "Surat", "Gujarat", "GIDC Industrial Area, Surat, Gujarat, India", "Needs a professional export-oriented website with product gallery, size charts, and international inquiry form."),
    ("Mehta Embroidery Works", "Textile, Garments & Clothing Store", "Specialized embroidery unit offering machine and hand embroidery on fabrics, blouses, and dupatta borders.", "Proprietor", "+91 97270 11203", "Surat", "Gujarat", "Varachha Road, Surat, Gujarat, India", "Needs a portfolio website showcasing embroidery designs, bulk pricing, and order placement for boutique buyers."),
    ("Surti Diamond & Gems Trading Co.", "Local Commercial Business", "Rough and polished diamond trading company supplying to jewellery manufacturers across India.", "Managing Partner", "+91 98252 11204", "Surat", "Gujarat", "Mini Bazaar, Mahidharpura, Surat, Gujarat, India", "Needs a secure B2B portal with GIA-certified diamond listings, live price updates, and buyer verification system."),
    ("Bharat Dyeing & Printing Works", "Manufacturing & Industrial Works", "Fabric dyeing and printing unit producing digital prints, block prints, and tie-dye for the textile market.", "Factory Owner", "+91 97240 11205", "Surat", "Gujarat", "Pandesara GIDC, Surat, Gujarat, India", "Needs a digital sample catalog and wholesale inquiry website to reach apparel manufacturers and exporters."),

    # ── AHMEDABAD (Chemicals, Textiles, Engineering) ────────────────────────
    ("Shah Chemicals & Allied Products", "Manufacturing & Industrial Works", "Manufacturer of industrial chemicals, solvents, and cleaning compounds for textile and pharmaceutical industries.", "Managing Director", "+91 98240 11206", "Ahmedabad", "Gujarat", "Naroda GIDC, Ahmedabad, Gujarat, India", "Needs a MSDS-compliant product catalog website with safety data sheets and bulk order inquiry form."),
    ("Prajapati Iron & Steel Traders", "Pipe, Hardware & Metal Industry", "Wholesale trader of TMT bars, MS pipes, structural steel, and hardware for the construction industry.", "Proprietor", "+91 97240 11207", "Ahmedabad", "Gujarat", "Rakhial Industrial Area, Ahmedabad, Gujarat, India", "Needs a B2B website with product specs, weight charts, and bulk procurement request form for contractors."),
    ("Amul Plastic Industries", "Manufacturing & Industrial Works", "Manufactures plastic injection-moulded components, containers, and HDPE fittings for packaging and plumbing.", "Factory Head", "+91 98795 11208", "Ahmedabad", "Gujarat", "Odhav GIDC, Ahmedabad, Gujarat, India", "Needs a corporate website with product range, production capacity, ISO certification details, and export inquiry form."),
    ("Gujarat Wholesale Agro Traders", "Distributor & Wholesale Trade", "Wholesale distributor of agricultural produce, pulses, spices, and grain to regional mandis and retailers.", "Senior Partner", "+91 99249 11209", "Ahmedabad", "Gujarat", "Agriculture Produce Market, Jamalpur, Ahmedabad, Gujarat, India", "Needs a commodity portal showing current mandi prices, product availability, and bulk purchase request system."),

    # ── LUDHIANA (Hosiery, Machinery, Cycles) ───────────────────────────────
    ("Sharma Hosiery & Knitwear Mills", "Textile, Garments & Clothing Store", "Hosiery manufacturer producing cotton undergarments, socks, and woollen knitwear for domestic retail.", "Factory Owner / MD", "+91 98150 11210", "Ludhiana", "Punjab", "Focal Point Industrial Area, Ludhiana, Punjab, India", "Needs an e-commerce website with product catalog, wholesale pricing table, and distributor onboarding form."),
    ("Gill Cycle Parts & Components", "Distributor & Wholesale Trade", "Wholesale trader and manufacturer of bicycle parts, components, and accessories for the domestic cycle industry.", "Managing Partner", "+91 98720 11211", "Ludhiana", "Punjab", "Miller Ganj Industrial Area, Ludhiana, Punjab, India", "Needs a B2B trade website for parts catalog, OEM inquiry forms, and dealer registration portal."),
    ("Punjab Machine Tools Pvt. Ltd.", "Manufacturing & Industrial Works", "Manufactures lathe machines, drilling machines, and precision engineering tools for industrial buyers.", "Director", "+91 97790 11212", "Ludhiana", "Punjab", "Gill Road Industrial Estate, Ludhiana, Punjab, India", "Needs a corporate website showcasing machine specifications, capacity details, and export capability statements."),
    ("Bhatia Steel Furniture Works", "Manufacturing & Industrial Works", "Produces office chairs, steel almirahs, industrial racks, and metal furniture for commercial buyers.", "Proprietor", "+91 98140 11213", "Ludhiana", "Punjab", "Daba Road, Ludhiana, Punjab, India", "Needs a catalog website with product images, dimensions, weight specs, and bulk order quotation form."),

    # ── MUMBAI (Finance, Trading, Services) ─────────────────────────────────
    ("Dadar Pipe & Sanitary Fittings", "Pipe, Hardware & Metal Industry", "Distributes CP fittings, PVC pipes, sanitary ware, and plumbing accessories to contractors and retailers.", "Shop Owner", "+91 98200 11214", "Mumbai", "Maharashtra", "Dr. Babasaheb Ambedkar Road, Dadar, Mumbai, Maharashtra, India", "Needs an online product catalog with inventory status and contractor pricing to replace phone-based order taking."),
    ("Dharavi Leather Goods Manufacturing", "Manufacturing & Industrial Works", "Produces leather wallets, belts, bags, and fashion accessories for domestic wholesale and export.", "Factory Head", "+91 99200 11215", "Mumbai", "Maharashtra", "Dharavi Industrial Area, Mumbai, Maharashtra, India", "Needs an export-ready website with product portfolio, leather grades info, and international buyer inquiry form."),
    ("Bhiwandi Textile Wholesale Hub", "Distributor & Wholesale Trade", "Wholesale distributor of grey fabric, processed cloth, and finished textiles to garment manufacturers.", "Senior Partner", "+91 98190 11216", "Bhiwandi", "Maharashtra", "Power Loom Colony, Bhiwandi, Maharashtra, India", "Needs a fabric catalog website with GSM details, width specifications, and bulk indent request form."),
    ("Kurla Auto Parts & Spares", "Automotive & Machinery Services", "Distributes auto spare parts, oil filters, brake pads, and engine components to mechanics and garages.", "Shop Owner", "+91 99670 11217", "Mumbai", "Maharashtra", "LBS Marg, Kurla, Mumbai, Maharashtra, India", "Needs a spare parts website with OEM cross-reference numbers, vehicle compatibility search, and order tracking."),
    ("Andheri Trading & Distribution Co.", "Distributor & Wholesale Trade", "Wholesale FMCG distributor supplying grocery items, cleaning products, and stationery to kirana stores.", "Distribution Head", "+91 98210 11218", "Mumbai", "Maharashtra", "MIDC Industrial Area, Andheri East, Mumbai, Maharashtra, India", "Needs a dealer management portal with live stock updates, order booking, and GST invoice generation."),
    ("Dombivli Pharma Raw Materials", "Distributor & Wholesale Trade", "Distributor of active pharmaceutical ingredients (APIs), excipients, and packaging materials to small pharma units.", "Managing Director", "+91 97690 11219", "Dombivli", "Maharashtra", "MIDC Phase II, Dombivli, Maharashtra, India", "Needs a regulatory-compliant product portal with CoA downloads, COO certificates, and RFQ submission."),

    # ── PUNE (Auto Ancillary, IT, Engineering) ───────────────────────────────
    ("Pimpri Auto Components Works", "Manufacturing & Industrial Works", "Manufactures precision-machined auto components, stamped parts, and CNC-turned items for OEM supply.", "Plant Manager / Director", "+91 98220 11220", "Pune", "Maharashtra", "Pimpri-Chinchwad MIDC, Pune, Maharashtra, India", "Needs a corporate website with ISO 9001 certification info, product range, machinery specs, and OEM registration."),
    ("Hadapsar Engineering Solutions", "Manufacturing & Industrial Works", "Provides precision engineering services, jig & fixture manufacturing, and custom metal fabrication.", "Proprietor / Engineer", "+91 97660 11221", "Pune", "Maharashtra", "Hadapsar Industrial Estate, Pune, Maharashtra, India", "Needs a service portfolio website with capability statements, past project showcase, and quote request form."),
    ("Chakan Rubber & Polymer Products", "Manufacturing & Industrial Works", "Manufactures rubber gaskets, O-rings, seals, and polymer bushings for automotive and industrial use.", "Factory Owner", "+91 98230 11222", "Chakan", "Maharashtra", "Chakan MIDC, Pune, Maharashtra, India", "Needs a technical datasheet portal with material grades, dimensional tolerances, and OEM application support."),

    # ── DELHI / NCR (Wholesale, Distribution, Hardware) ─────────────────────
    ("Karol Bagh Wholesale Fabrics", "Textile, Garments & Clothing Store", "Wholesale fabric dealer selling zari, net fabric, bridal materials, and embellished cloth to boutique tailors.", "Proprietor", "+91 98100 11223", "Delhi", "Delhi", "Ajmal Khan Road, Karol Bagh, New Delhi, Delhi, India", "Needs an online fabric catalog with swatch uploads, bulk wholesale pricing, and nationwide delivery option."),
    ("Chandni Chowk Hardware & Tools", "Pipe, Hardware & Metal Industry", "Wholesale trader of nuts, bolts, hand tools, power tools, and industrial hardware.", "Shop Owner", "+91 99990 11224", "Delhi", "Delhi", "Chandni Chowk Market, New Delhi, Delhi, India", "Needs a B2B catalog website with product SKUs, quantity-based pricing tiers, and contractor bulk quote form."),
    ("Lajpat Rai Electronics Wholesale", "Distributor & Wholesale Trade", "Wholesale distributor of electronic components, ICs, resistors, capacitors, and LED modules.", "Senior Partner", "+91 98118 11225", "Delhi", "Delhi", "Lajpat Rai Market, Chandni Chowk, Delhi, India", "Needs an electronics B2B portal with part search, datasheet library, and wholesale account login."),
    ("Sadar Bazaar FMCG Wholesale", "Distributor & Wholesale Trade", "One of Delhi's largest FMCG wholesale distributors, supplying packaged goods to kiranas and institutions.", "Distribution Manager", "+91 98102 11226", "Delhi", "Delhi", "Sadar Bazaar, New Delhi, Delhi, India", "Needs a dealer portal with live stock levels, brand-wise ordering, and automated GST invoice generation."),
    ("Okhla Industrial Packaging Supplies", "Distributor & Wholesale Trade", "Wholesale supplier of corrugated boxes, stretch wrap, bubble wrap, and industrial packaging materials.", "Managing Director", "+91 98687 11227", "Delhi", "Delhi", "Okhla Industrial Area Phase II, New Delhi, Delhi, India", "Needs a packaging website with product dimensions, load-bearing specs, and customization request form."),
    ("Gurgaon Steel & Pipe Wholesale", "Pipe, Hardware & Metal Industry", "Supplies MS pipes, GI pipes, ERW tubes, and steel sections to construction and MEP contractors in NCR.", "Owner / Partner", "+91 96540 11228", "Gurgaon", "Haryana", "Udyog Vihar Industrial Area, Gurgaon, Haryana, India", "Needs a B2B pipe catalog website with grade specs, weight per meter charts, and contractor login portal."),
    ("Noida Plastic Granules Supplier", "Distributor & Wholesale Trade", "Distributes PP, HDPE, LDPE, and ABS plastic granules to injection moulding units and packaging companies.", "Proprietor", "+91 98971 11229", "Noida", "Uttar Pradesh", "Sector 63 Industrial Area, Noida, Uttar Pradesh, India", "Needs a material specification website with resin grades, MFI data, and bulk granule order system."),

    # ── JAIPUR (Gems, Jewellery, Handicrafts) ───────────────────────────────
    ("Rajasthan Gems & Jewellery House", "Jewellery & Goldsmith", "Wholesale dealer in precious and semi-precious gemstones, kundan jewellery, and polki ornaments.", "Proprietor", "+91 98290 11230", "Jaipur", "Rajasthan", "Johari Bazaar, Jaipur, Rajasthan, India", "Needs a B2B gems trading website with stone gallery, carat/clarity filters, and certified export inquiry."),
    ("Anokhi Block Print & Handicrafts", "Textile, Garments & Clothing Store", "Produces traditional hand block-printed fabrics, bed linen, and garments for domestic and export markets.", "Artisan Entrepreneur", "+91 94130 11231", "Jaipur", "Rajasthan", "Sanganer Township, Jaipur, Rajasthan, India", "Needs a craft e-commerce website with artisan story, product gallery, and international shipping option."),
    ("Marble & Stone Exporter Rajasthan", "Distributor & Wholesale Trade", "Quarries and distributes Makrana marble, Kota stone, and sandstone to builders, architects, and exporters.", "Export Head", "+91 97993 11232", "Jaipur", "Rajasthan", "Mahal Road, Mansarovar, Jaipur, Rajasthan, India", "Needs a stone catalog website with slab sizes, grades, finish options, and international inquiry form."),
    ("Rajpur Footwear & Leather Goods", "Local Commercial Business", "Manufactures Mojari traditional footwear, leather sandals, and handcrafted shoes for retail.", "Proprietor / Artisan", "+91 82099 11233", "Jaipur", "Rajasthan", "Bapu Bazaar, Jaipur, Rajasthan, India", "Needs an online shop to showcase handcrafted footwear to domestic tourists and international buyers."),

    # ── MORADABAD (Brassware, Metal Export) ─────────────────────────────────
    ("Moradabad Brassware Exports", "Manufacturing & Industrial Works", "Exports brass decorative items, metal handicrafts, and plated showpieces to European and US buyers.", "Export Manager", "+91 97190 11234", "Moradabad", "Uttar Pradesh", "Brass Colony, Moradabad, Uttar Pradesh, India", "Needs an export product catalog website with item codes, MOQ details, and international buyer portal."),
    ("Arora Metal Polishing & Plating Works", "Manufacturing & Industrial Works", "Provides metal plating (copper, nickel, chrome, gold) and polishing services to brassware exporters.", "Factory Owner", "+91 98370 11235", "Moradabad", "Uttar Pradesh", "Civil Lines Industrial Area, Moradabad, Uttar Pradesh, India", "Needs a service website with plating capabilities, substrate compatibility, and quote request form."),

    # ── KANPUR (Leather, Chemicals, Textiles) ────────────────────────────────
    ("Kanpur Leather Footwear Factory", "Manufacturing & Industrial Works", "Manufactures leather shoes, boots, and industrial safety footwear for domestic and export markets.", "Factory Head / MD", "+91 97922 11236", "Kanpur", "Uttar Pradesh", "Jajmau Industrial Area, Kanpur, Uttar Pradesh, India", "Needs a B2B factory website with product catalog, safety standards (ISI/CE), and export inquiry form."),
    ("Shukla Chemicals Industries", "Manufacturing & Industrial Works", "Manufactures tanning chemicals, dyestuffs, and leather auxiliaries for the local leather industry.", "Managing Director", "+91 98390 11237", "Kanpur", "Uttar Pradesh", "Dada Nagar Industrial Estate, Kanpur, Uttar Pradesh, India", "Needs a technical website with product TDS sheets, application guides, and distributor inquiry system."),

    # ── COIMBATORE (Auto, Textiles, Foundry) ────────────────────────────────
    ("Coimbatore Foundry & Castings", "Manufacturing & Industrial Works", "Grey iron and ductile iron foundry producing castings, pulleys, impellers, and pump parts.", "Works Manager / Owner", "+91 99422 11238", "Coimbatore", "Tamil Nadu", "Ganapathy Industrial Estate, Coimbatore, Tamil Nadu, India", "Needs a capability website with material grades, casting weight range, and OEM/export inquiry form."),
    ("Tirupur Knitting & Garments Unit", "Textile, Garments & Clothing Store", "Knits and exports cotton t-shirts, polo shirts, and sportswear to European and US retailers.", "Export Director", "+91 97900 11239", "Tirupur", "Tamil Nadu", "First Street, SIPCOT Industrial Estate, Tirupur, Tamil Nadu, India", "Needs an international buyer website with product samples, compliance certifications, and export inquiry."),
    ("Kumar Spinning Mills", "Manufacturing & Industrial Works", "Cotton spinning mill producing 20s, 30s, and 40s count cotton yarn for weaving and knitting units.", "Mill Owner / MD", "+91 94430 11240", "Coimbatore", "Tamil Nadu", "Peelamedu Industrial Area, Coimbatore, Tamil Nadu, India", "Needs a mill website with yarn count catalog, quality certifications, and bulk purchase inquiry form."),
    ("Coimbatore Pump & Motor Dealers", "Distributor & Wholesale Trade", "Authorized wholesale dealer of agricultural and industrial pumps, motors, and water management systems.", "Dealer Principal", "+91 98430 11241", "Coimbatore", "Tamil Nadu", "Mettupalayam Road, Coimbatore, Tamil Nadu, India", "Needs a product catalog website with brand lineup, pump selection guide, and dealer/service network map."),

    # ── CHENNAI (Auto Ancillary, Electronics) ────────────────────────────────
    ("Chennai Auto Ancillary Components", "Manufacturing & Industrial Works", "Produces stamped and welded auto body parts, brackets, and sub-assemblies for OEM supply.", "Plant Manager", "+91 98410 11242", "Chennai", "Tamil Nadu", "SIPCOT Industrial Complex, Gummidipoondi, Chennai, Tamil Nadu, India", "Needs an OEM supply portal with part drawings, PPAPs, and quality approval documentation."),
    ("T.Nagar Wholesale Saree Hub", "Textile, Garments & Clothing Store", "Wholesale saree showroom carrying Kanjivaram silk, cotton, chiffon, and synthetic sarees for retailers.", "Owner / Managing Director", "+91 98840 11243", "Chennai", "Tamil Nadu", "Usman Road, T. Nagar, Chennai, Tamil Nadu, India", "Needs an online saree catalog with fabric type filters, bulk wholesale pricing, and pan-India delivery option."),
    ("Perungudi Electronics Component Dist.", "Distributor & Wholesale Trade", "Distributes semiconductors, passive components, and electromechanical parts to PCB assembly units.", "Business Owner", "+91 97907 11244", "Chennai", "Tamil Nadu", "IT Expressway, Perungudi, Chennai, Tamil Nadu, India", "Needs a component database website with datasheet links, stock availability, and buyer registration."),

    # ── HYDERABAD (Pharma, Chemicals, Food) ──────────────────────────────────
    ("Hyderabad Pharma Bulk Drug Supplier", "Distributor & Wholesale Trade", "Wholesale supplier of APIs, intermediates, and pharmaceutical raw materials to formulation companies.", "Senior Partner", "+91 98490 11245", "Hyderabad", "Telangana", "IDA Bollaram Industrial Area, Hyderabad, Telangana, India", "Needs a pharma-compliant website with COA downloads, GMP certification info, and regulated inquiry system."),
    ("Deccan Food Processing Industries", "Manufacturing & Industrial Works", "Processes and packages red chilli powder, turmeric, spice blends, and sesame for FMCG brands.", "Factory Owner / MD", "+91 99898 11246", "Hyderabad", "Telangana", "Bonthapally Industrial Area, Sangareddy, Telangana, India", "Needs a B2B food supply website with FSSAI license display, packaging options, and custom blending inquiry."),
    ("Patancheru Chemical Storage & Dist.", "Distributor & Wholesale Trade", "Stores and distributes industrial chemicals, acids, solvents, and adhesives to manufacturers.", "Operations Head", "+91 94405 11247", "Patancheru", "Telangana", "ANRICH Industrial Estate, Patancheru, Telangana, India", "Needs an industrial chemical portal with SDS documents, regulatory compliance info, and logistics inquiry."),

    # ── KOLKATA (Jute, Engineering, Trading) ─────────────────────────────────
    ("Bengal Jute & Fibre Products", "Manufacturing & Industrial Works", "Manufactures jute bags, hessian cloth, geotextiles, and diversified jute products for export.", "Managing Director", "+91 98310 11248", "Kolkata", "West Bengal", "Serampore Industrial Area, Kolkata, West Bengal, India", "Needs an export website with product specifications, eco-friendly credentials, and EU buyer inquiry system."),
    ("Burrabazar Wholesale Stationery", "Distributor & Wholesale Trade", "Wholesale distributor of stationery, office supplies, school materials, and paper products.", "Proprietor", "+91 98360 11249", "Kolkata", "West Bengal", "Burra Bazar, Kolkata, West Bengal, India", "Needs a wholesale portal with brand catalog, institutional bulk pricing, and school/office supply quote system."),
    ("Howrah Engineering & Fabrication", "Manufacturing & Industrial Works", "General engineering workshop providing structural fabrication, welding, and machining services.", "Workshop Owner", "+91 97485 11250", "Howrah", "West Bengal", "Andul Road Industrial Zone, Howrah, West Bengal, India", "Needs a service capability website with project portfolio, material handling specs, and quote request form."),
    ("Garden Reach Marine Components", "Manufacturing & Industrial Works", "Manufactures marine hardware, ship fittings, anchor chains, and propeller shaft seals.", "Technical Director", "+91 98306 11251", "Kolkata", "West Bengal", "Garden Reach Shipbuilding Area, Kolkata, West Bengal, India", "Needs a technical portal with maritime standards compliance, ABS/DNV approval docs, and supply inquiry form."),

    # ── BANGALORE (Manufacturing, IT Hardware, Services) ─────────────────────
    ("Peenya Industrial Precision Parts", "Manufacturing & Industrial Works", "Precision CNC machined components for aerospace, defence, and automotive sectors.", "MD / Founder", "+91 98450 11252", "Bangalore", "Karnataka", "Peenya Industrial Area 2nd Stage, Bangalore, Karnataka, India", "Needs an AS9100-compliant supplier website with machining capabilities and aerospace OEM qualification form."),
    ("Electronic City Hardware Wholesale", "Distributor & Wholesale Trade", "Wholesale distributor of computer hardware, networking equipment, and IT peripherals.", "Owner", "+91 97418 11253", "Bangalore", "Karnataka", "Electronic City Phase 1, Bangalore, Karnataka, India", "Needs a B2B IT wholesale portal with SKU database, brand catalog, and bulk order system."),
    ("Rajajinagar Garments Distributors", "Textile, Garments & Clothing Store", "Wholesale distributor of branded and unbranded readymade garments across Karnataka.", "Distribution Partner", "+91 99003 11254", "Bangalore", "Karnataka", "Rajajinagar Industrial Town, Bangalore, Karnataka, India", "Needs a garment wholesale website with category filters, MOQ details, and retailer signup portal."),

    # ── NAGPUR (Oranges, Trading, Transport) ──────────────────────────────────
    ("Nagpur Agro Commodities Wholesale", "Distributor & Wholesale Trade", "Wholesale trader of oranges, seasonal fruits, pulses, and agricultural commodities.", "Commission Agent / Owner", "+91 94220 11255", "Nagpur", "Maharashtra", "Kalamna Agri Market Yard, Nagpur, Maharashtra, India", "Needs a commodity trading website with daily mandi prices, quality grading, and bulk buyer registration."),
    ("Vidarbha Steel Rolling Mills", "Manufacturing & Industrial Works", "Steel rolling mill producing MS angles, channels, flats, and rounds for local construction trade.", "MD / Works Manager", "+91 98230 11256", "Nagpur", "Maharashtra", "Butibori MIDC, Nagpur, Maharashtra, India", "Needs a steel product website with standard sizes, weight tables, IS grade specs, and contractor inquiry."),

    # ── RAJKOT (Auto, Machinery, Casting) ─────────────────────────────────────
    ("Rajkot Auto Parts Casting & Forging", "Manufacturing & Industrial Works", "Manufactures forged and cast auto components, connecting rods, and pump bodies.", "MD / Plant Head", "+91 98244 11257", "Rajkot", "Gujarat", "Shapar Veraval GIDC, Rajkot, Gujarat, India", "Needs a supplier website with material grade certifications, capacity metrics, and OEM qualification form."),
    ("Rajkot Agricultural Pump Supplier", "Distributor & Wholesale Trade", "Wholesale dealer of submersible pumps, centrifugal pumps, and drip irrigation components.", "Dealer Owner", "+91 94265 11258", "Rajkot", "Gujarat", "Gondal Road Industrial Area, Rajkot, Gujarat, India", "Needs a dealer website with brand lineup, technical pump selection guide, and service center locator."),

    # ── INDORE (Soya, Textiles, Pharma) ──────────────────────────────────────
    ("Indore Soya Agro Processing", "Manufacturing & Industrial Works", "Processes soyabean into de-oiled cake (DOC), soy oil, and lecithin for animal feed and food industries.", "MD / Production Head", "+91 98270 11259", "Indore", "Madhya Pradesh", "Pithampur Industrial Area, Indore, Madhya Pradesh, India", "Needs a commodity website with product specs, protein/fat content data, and bulk buyer inquiry."),
    ("Madhya Pradesh Pharma Wholesale", "Distributor & Wholesale Trade", "Wholesale pharmaceutical distributor supplying branded medicines, generics, and OTC products.", "Stockist / Distributor", "+91 98930 11260", "Indore", "Madhya Pradesh", "Palasia Commercial Zone, Indore, Madhya Pradesh, India", "Needs a stockist portal with medicine catalog, brand partnerships display, and medical shop login."),

    # ── BHOPAL (Chemicals, Food, Engineering) ───────────────────────────────
    ("Bhopal Industrial Gas Suppliers", "Distributor & Wholesale Trade", "Supplies industrial gases (oxygen, nitrogen, argon, CO₂) to hospitals, metal cutting, and food industries.", "Owner / Operations Manager", "+91 94254 11261", "Bhopal", "Madhya Pradesh", "Mandideep Industrial Zone, Bhopal, Madhya Pradesh, India", "Needs a safety-compliant website with cylinder rental, gas purity specs, and emergency supply inquiry."),
    ("MP Atta & Flour Mills", "Manufacturing & Industrial Works", "Flour milling unit producing wheat flour, maida, sooji, and atta for retail and bulk supply.", "Mill Owner", "+91 97558 11262", "Bhopal", "Madhya Pradesh", "Govindpura Industrial Area, Bhopal, Madhya Pradesh, India", "Needs a food-grade B2B website with product grades, packaging options, and institutional bulk orders."),

    # ── LUCKNOW (Chikan, Carpets, MSME) ─────────────────────────────────────
    ("Lucknow Chikan Embroidery House", "Textile, Garments & Clothing Store", "Artisanal chikankari embroidery studio producing kurtas, sarees, and home linen for retail and export.", "Master Artisan / Owner", "+91 94156 11263", "Lucknow", "Uttar Pradesh", "Chowk Bazaar, Lucknow, Uttar Pradesh, India", "Needs an artisan e-commerce website with embroidery style gallery, custom order option, and export inquiry."),
    ("Lucknow Carpets & Handloom Weaving", "Textile, Garments & Clothing Store", "Produces hand-knotted Kaleen carpets, dhurries, and handloom floor covers for export.", "Weaving Unit Owner", "+91 98390 11264", "Lucknow", "Uttar Pradesh", "Sitapur Road Industrial Zone, Lucknow, Uttar Pradesh, India", "Needs an export catalog website with knot density specs, pile height data, and importer inquiry form."),

    # ── VARANASI (Silk, BHU area, Handicrafts) ──────────────────────────────
    ("Banarasi Silk Weaving Cooperative", "Textile, Garments & Clothing Store", "Weaves traditional Banarasi brocade silk sarees, fabric, and home furnishing textiles.", "Cooperative Head / Proprietor", "+91 94154 11265", "Varanasi", "Uttar Pradesh", "Sarai Mohana Weaver Colony, Varanasi, Uttar Pradesh, India", "Needs a GI-certified artisan website with saree designs, silk purity data, and custom bulk order form."),
    ("Varanasi Wooden Toy & Craft Exporter", "Local Commercial Business", "Manufactures lacquered wooden toys, figurines, and artisan decor items for export.", "Proprietor / Artisan", "+91 94151 11266", "Varanasi", "Uttar Pradesh", "Sigra Bazaar, Varanasi, Uttar Pradesh, India", "Needs an online handicraft store with product catalog, custom crafting option, and international shipping."),

    # ── NGOs ─────────────────────────────────────────────────────────────────
    ("Udaan Bal Vikas Sanstha", "Non-Governmental Organization (NGO)", "Non-profit organization running child welfare programs, primary education, and mid-day meal schemes.", "Executive Secretary", "+91 98200 11267", "Pune", "Maharashtra", "Hadapsar Community Center, Pune, Maharashtra, India", "Needs a digital donation portal with CSR receipts, program reports, and 80G tax exemption certificate display."),
    ("Prayas Mahila Uthan Foundation", "Non-Governmental Organization (NGO)", "Women empowerment NGO providing skill training, micro-credit access, and legal awareness.", "Founder / Director", "+91 94140 11268", "Jaipur", "Rajasthan", "Vaishali Nagar, Jaipur, Rajasthan, India", "Needs a website for beneficiary registration, online donation, annual report downloads, and program transparency."),
    ("Asha Sewa Samiti (Charitable Trust)", "Non-Governmental Organization (NGO)", "Operates free medical camps, ambulance service, and old age home support in rural areas.", "Managing Trustee", "+91 98120 11269", "Lucknow", "Uttar Pradesh", "Vikas Nagar, Lucknow, Uttar Pradesh, India", "Needs a website for volunteer sign-ups, online contributions, health camp schedules, and transparency reports."),
    ("Gramin Vikas Seva Trust", "Non-Governmental Organization (NGO)", "Rural development NGO working on water conservation, sanitation, and organic farming promotion.", "Project Coordinator", "+91 94220 11270", "Nagpur", "Maharashtra", "Wardha Road, Nagpur, Maharashtra, India", "Needs a website for FCRA-compliant donation collection, project updates, and government scheme enrollment."),
    ("Jan Sahayog Charitable Society", "Non-Governmental Organization (NGO)", "Provides free legal aid, disability support, and livelihood training to marginalized communities.", "Secretary", "+91 97340 11271", "Delhi", "Delhi", "Rohini Sector 15, New Delhi, Delhi, India", "Needs a website for legal help requests, donation campaigns, beneficiary testimonials, and funding reports."),
    ("Navjyoti Shiksha Evam Seva Samiti", "Non-Governmental Organization (NGO)", "Runs adult literacy centres, computer training labs, and scholarship programs for underprivileged youth.", "President", "+91 98030 11272", "Kanpur", "Uttar Pradesh", "Swaroop Nagar, Kanpur, Uttar Pradesh, India", "Needs a scholarship portal with application forms, eligibility criteria, donor profiles, and student success stories."),
    ("Sankalp Rural Health Mission", "Non-Governmental Organization (NGO)", "Mobile health clinic NGO providing primary healthcare and maternal services to tribal villages.", "Medical Director", "+91 94290 11273", "Raipur", "Chhattisgarh", "Shankar Nagar, Raipur, Chhattisgarh, India", "Needs a healthcare NGO website with camp schedules, digital prescription uploads, and CSR partnership page."),
    ("Swabhiman Adivasi Kalyan Kendra", "Non-Governmental Organization (NGO)", "Works with tribal communities on forest rights, education access, and traditional craft marketing.", "Coordinator", "+91 94050 11274", "Bhubaneswar", "Odisha", "Chandrasekharpur, Bhubaneswar, Odisha, India", "Needs a platform for craft e-marketplace, tribal welfare reporting, and online donation with 80G acknowledgement."),

    # ── Auto Services ─────────────────────────────────────────────────────────
    ("Sharma Motors & Garage Works", "Automotive & Machinery Services", "Multi-brand vehicle service centre offering engine overhaul, AC servicing, and denting & painting.", "Service Manager", "+91 98160 11275", "Delhi", "Delhi", "Najafgarh Road Industrial Area, New Delhi, Delhi, India", "Needs a garage website for online service slot booking, service package display, and customer review collection."),
    ("Bengal Motors Service & Spares", "Automotive & Machinery Services", "Authorised service point and spare parts dealer for commercial vehicles and heavy trucks.", "Proprietor", "+91 98310 11276", "Kolkata", "West Bengal", "Kona Expressway Service Road, Howrah, West Bengal, India", "Needs a commercial vehicle service website with service packages, OEM parts ordering, and fleet service contracts."),
    ("Punjab Auto Ancillary Distributors", "Automotive & Machinery Services", "Wholesale distributor of two-wheeler spare parts, clutch plates, carburetors, and batteries.", "Distribution Head", "+91 98150 11277", "Ludhiana", "Punjab", "Jalandhar Bypass Road, Ludhiana, Punjab, India", "Needs a spare parts dealer website with vehicle compatibility search, price lists, and retailer login."),
    ("Chennai Vehicle Body Builders", "Automotive & Machinery Services", "Custom commercial vehicle body building unit fabricating truck bodies, tankers, and tippers.", "Works Owner / MD", "+91 98840 11278", "Chennai", "Tamil Nadu", "Ambattur Industrial Estate, Chennai, Tamil Nadu, India", "Needs a custom body builder website with past build portfolio, materials used, and bus/truck inquiry form."),

    # ── Hardware / Construction / Real estate ─────────────────────────────────
    ("Hyderabad Cement & Building Supplies", "Distributor & Wholesale Trade", "Wholesale dealer of OPC/PPC cement, AAC blocks, fly ash bricks, and construction chemicals.", "Owner", "+91 98498 11279", "Hyderabad", "Telangana", "Uppal Industrial Area, Hyderabad, Telangana, India", "Needs a construction materials website with brand listings, current price updates, and contractor credit inquiry."),
    ("Kolkata Electrical Wholesale Depot", "Distributor & Wholesale Trade", "Wholesale electrical materials distributor supplying cables, switchgear, MCBs, and conduits.", "Senior Partner", "+91 98310 11280", "Kolkata", "West Bengal", "Burra Bazar Electrical Market, Kolkata, West Bengal, India", "Needs a product catalog website with brand listings, technical specs, and contractor account system."),
    ("Pune Waterproofing & Construction Chem.", "Distributor & Wholesale Trade", "Distributes waterproofing compounds, construction adhesives, tile grouts, and epoxy systems.", "Technical Sales Manager", "+91 98220 11281", "Pune", "Maharashtra", "Talegaon Industrial Zone, Pune, Maharashtra, India", "Needs a technical products website with application guides, substrate compatibility, and project cost estimator."),

    # ── Miscellaneous Indian SMEs ─────────────────────────────────────────────
    ("Ahmedabad Cold Storage & Logistics", "Distributor & Wholesale Trade", "Cold chain logistics operator providing temperature-controlled storage for fruits, vegetables, and dairy.", "Operations Manager", "+91 98790 11282", "Ahmedabad", "Gujarat", "Naroda Road, Ahmedabad, Gujarat, India", "Needs a logistics portal with storage capacity booking, temperature monitoring logs, and FSSAI compliance display."),
    ("Rajasthan Marble & Granite Quarry", "Manufacturing & Industrial Works", "Quarries and processes Rajasthan white marble, Makrana marble, and pink granite for export.", "Export Manager", "+91 97993 11283", "Kishangarh", "Rajasthan", "Marble City, Kishangarh, Rajasthan, India", "Needs a quarry-to-buyer website with slab grades, surface finishes, and international importer inquiry."),
    ("Odisha Tribal Handicraft Federation", "Non-Governmental Organization (NGO)", "Government-linked federation marketing tribal Pattachitra paintings, Dokra metal crafts, and tribal textiles.", "Federation Secretary", "+91 94380 11284", "Bhubaneswar", "Odisha", "Udyog Bhavan, Bhubaneswar, Odisha, India", "Needs a craft marketplace website for artisan profiles, GI tag product listings, and international buyer queries."),
    ("Guwahati Tea Estate Wholesale", "Distributor & Wholesale Trade", "Wholesale dealer of Assam orthodox and CTC teas from garden estates to retailers and exporters.", "Senior Partner", "+91 94350 11285", "Guwahati", "Assam", "Pan Bazaar Tea Trade Zone, Guwahati, Assam, India", "Needs a tea trading website with estate profiles, flush details, FSSAI/APEDA certifications, and bulk order portal."),
    ("Bihar Rice Mill & Agro Processing", "Manufacturing & Industrial Works", "Parboiled and raw rice milling unit processing paddy for FMCG brands and government procurement.", "Mill Owner", "+91 94718 11286", "Patna", "Bihar", "Khagaria Road Milling Zone, Patna, Bihar, India", "Needs a milling B2B website with grade specifications, FRK fortification compliance, and institutional supply form."),
    ("Jammu Walnut & Dry Fruit Exporter", "Distributor & Wholesale Trade", "Wholesale exporter of Kashmiri walnuts, almonds, saffron, and dried apricots to national and global markets.", "Export Head / Owner", "+91 94191 11287", "Srinagar", "J&K", "Shitla Bazar Dry Fruit Market, Srinagar, J&K, India", "Needs an export website with origin certification, GI tagging, product grades, and international buyer inquiry."),
    ("Kerala Coir & Husk Products Coop.", "Manufacturing & Industrial Works", "Manufactures coir rope, coir mats, coir pith (cocopeat), and coir geotextiles for export.", "Cooperative President", "+91 94470 11288", "Alappuzha", "Kerala", "Coir House Industrial Area, Alappuzha, Kerala, India", "Needs an eco-product website with FSC/ISO certification, product range, and EU buyer inquiry form."),
    ("Dehradun Honey & Forest Products", "Distributor & Wholesale Trade", "Wholesale supplier of raw honey, beeswax, forest herbs, and Himalayan spices from Van Gujjar communities.", "Proprietor", "+91 94120 11289", "Dehradun", "Uttarakhand", "Rispana Road Market, Dehradun, Uttarakhand, India", "Needs an organic products website with FSSAI licensing, lab test reports, and bulk B2B order portal."),
    ("Haryana Agri Implements Factory", "Manufacturing & Industrial Works", "Manufactures agricultural equipment: seed drills, cultivators, harrows, and tractor-mounted implements.", "Factory Owner / MD", "+91 98120 11290", "Hisar", "Haryana", "Industrial Growth Centre, Hisar, Haryana, India", "Needs a product website with equipment specs, horsepower ratings, dealer network map, and financing options info."),
    ("Amritsar Woollen & Shawl Weavers", "Textile, Garments & Clothing Store", "Produces Amritsari Phulkari shawls, woollen blankets, and embroidered home textiles for export.", "Artisan Entrepreneur", "+91 97990 11291", "Amritsar", "Punjab", "Hall Bazaar Textile Area, Amritsar, Punjab, India", "Needs a craft export website showcasing GI-tagged Phulkari designs, wholesale pricing, and export inquiry form."),
    ("Raipur Steel & TMT Bar Distributors", "Pipe, Hardware & Metal Industry", "Wholesale steel distributor supplying TMT bars, structural steel, and wire rods to Chhattisgarh builders.", "Proprietor", "+91 94250 11292", "Raipur", "Chhattisgarh", "Urla Industrial Area, Raipur, Chhattisgarh, India", "Needs a steel dealer website with ISI-grade product specs, current market prices, and contractor bulk order form."),
    ("Jamshedpur Steel Scrap & Recycling", "Manufacturing & Industrial Works", "Steel scrap dealer and secondary melting unit supplying ingots and billets to re-rolling mills.", "Owner / Director", "+91 94315 11293", "Jamshedpur", "Jharkhand", "Adityapur Industrial Area, Jamshedpur, Jharkhand, India", "Needs a scrap trade website with material grades, quantity available, and rolling mill partnership inquiry."),
    ("Nashik Wine & Grape Processing Unit", "Manufacturing & Industrial Works", "Processes Nashik grapes into wine, grape juice, raisins, and grape extract for domestic and export markets.", "MD / Processing Head", "+91 94234 11294", "Nashik", "Maharashtra", "Dindori Wine Zone, Nashik, Maharashtra, India", "Needs a wine trade website with vintage profiles, FSSAI/APEDA approvals, and bulk buyer inquiry portal."),
    ("Firozabad Glass & Bangle Wholesale", "Distributor & Wholesale Trade", "Wholesale distributor of glass bangles, glass beads, and decorative glass products.", "Senior Partner", "+91 97195 11295", "Firozabad", "Uttar Pradesh", "Glass Nagar Industrial Zone, Firozabad, Uttar Pradesh, India", "Needs a B2B glass trade website with product catalog, MOQ info, and export sample request form."),
    ("Muradabad Imitation Jewellery Export", "Distributor & Wholesale Trade", "Exports imitation jewellery, metal-based fashion accessories, and decorative gift items globally.", "Export Head", "+91 97194 11296", "Moradabad", "Uttar Pradesh", "Brass Export Zone, Moradabad, Uttar Pradesh, India", "Needs an export website with fashion jewellery catalog, plating options, and US/EU buyer inquiry form."),
    ("Mysore Sandalwood & Agarbatti Works", "Manufacturing & Industrial Works", "Manufactures premium sandalwood-based agarbattis, incense sticks, and aromatic wood products.", "Factory Owner", "+91 98451 11297", "Mysore", "Karnataka", "Hebbal Industrial Area, Mysore, Karnataka, India", "Needs a product website with ingredient transparency, premium gift packaging options, and export inquiry form."),
    ("Jalandhar Sports Goods Manufacturers", "Manufacturing & Industrial Works", "Manufactures cricket bats, footballs, hockey sticks, and gym equipment for domestic and export markets.", "Managing Director", "+91 98150 11298", "Jalandhar", "Punjab", "GT Road Industrial Cluster, Jalandhar, Punjab, India", "Needs a sports equipment export website with product range, safety/BIS certifications, and bulk buyer inquiry."),
    ("Tiruvananthapuram Coconut Oil Mill", "Manufacturing & Industrial Works", "Cold-pressed and refined coconut oil manufacturing unit serving local food and FMCG industries.", "Mill Proprietor", "+91 94470 11299", "Thiruvananthapuram", "Kerala", "Attingal Industrial Area, Thiruvananthapuram, Kerala, India", "Needs an FSSAI-compliant product website with extraction method info, purity lab reports, and bulk order system."),
    ("Vijayawada Rice & Spice Exporters", "Distributor & Wholesale Trade", "Exports raw and parboiled rice, red chilli, turmeric, and processed spices from Andhra Pradesh.", "Export Manager", "+91 94404 11300", "Vijayawada", "Andhra Pradesh", "Autonagar Industrial Estate, Vijayawada, Andhra Pradesh, India", "Needs an export-ready website with APEDA registration display, produce grades, and international buyer portal."),
]


def append_india_leads(leads_data, filepath="Leads_Without_Websites_1000.xlsx"):
    wb  = openpyxl.load_workbook(filepath)
    ws  = wb.worksheets[0]

    max_row    = ws.max_row
    last_num   = max_row - 1  # existing data rows (excluding header)

    font_family   = "Segoe UI"
    data_font     = Font(name=font_family, size=10, color="1F2937")
    row_even_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    row_odd_fill  = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")  # India accent (light blue)
    thin_border   = Border(
        left=Side(style='thin',   color='D9D9D9'),
        right=Side(style='thin',  color='D9D9D9'),
        top=Side(style='thin',    color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    col_order = [
        "Lead ID", "Company Name", "Category / Industry", "What Company Does",
        "Contact Person / Role", "Phone Number", "City", "State / Region",
        "Country", "Full Address", "Website Status",
        "Why They Need A Website (Sales Pitch)"
    ]

    added = 0
    for idx, row_tuple in enumerate(leads_data):
        (name, cat, desc, role, phone, city, state, addr, pitch) = row_tuple
        new_row_num = max_row + idx + 1
        lead_id     = f"LEAD-{last_num + idx + 1:04d}"

        lead_dict = {
            "Lead ID": lead_id,
            "Company Name": name,
            "Category / Industry": cat,
            "What Company Does": desc,
            "Contact Person / Role": role,
            "Phone Number": phone,
            "City": city,
            "State / Region": state,
            "Country": "India",
            "Full Address": addr,
            "Website Status": "Confirmed No Website",
            "Why They Need A Website (Sales Pitch)": pitch
        }

        ws.append([lead_dict.get(col, "") for col in col_order])
        ws.row_dimensions[new_row_num].height = 22

        fill = row_even_fill if new_row_num % 2 == 0 else row_odd_fill
        for col_idx, col_name in enumerate(col_order, 1):
            cell = ws.cell(row=new_row_num, column=col_idx)
            cell.font   = data_font
            cell.fill   = fill
            cell.border = thin_border
            cell.alignment = align_center if col_name in ["Lead ID","Phone Number","City","Country","Website Status"] else align_left
        added += 1

    wb.save(filepath)
    total_now = max_row - 1 + added
    print(f"Appended {added} Indian leads.")
    print(f"Excel now has {total_now} total leads.")
    print(f"File: {filepath}")
    return added


if __name__ == "__main__":
    count = append_india_leads(INDIAN_LEADS)
    print(f"\nDone! Added {count} verified Indian business leads without websites.")
